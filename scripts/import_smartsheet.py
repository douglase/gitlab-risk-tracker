#!/usr/bin/env python3
"""Migrate Smartsheets risk text into GitLab issue descriptions.

Reads an .xlsx export with (at minimum) these columns:

  - ``GitLab Link``                 — full URL of the issue, e.g.
                                      https://gitlab.example.com/grp/proj/-/issues/123
  - ``Risk Description``            — body for the ``## Risk Description`` section
  - ``Action Plan/ Notes``          — body for the ``## Notes`` section
    (``Action Plan / Notes`` or ``Notes`` are also accepted)
  - ``Risk Mitigation Planning``    — body for the ``## Mitigation Plan`` section
    (``Mitigation Plan`` is also accepted)

For each row:

1. Parse the issue URL → ``server``, ``project_path``, ``iid``.
2. GET the current description via the GitLab REST API.
3. For each canonical section:

   - If the existing section's body matches the spreadsheet cell
     (modulo trailing whitespace / blank-line runs), do nothing.
   - If they differ, OR if the issue has no such section yet, append
     a clearly-marked **proposal block** at the end of the description
     containing the spreadsheet text and a unified diff against the
     current section. **The existing canonical section is never
     modified.**
4. If the new description differs from the old, PUT it back.

Designed to be safely re-runnable: each proposal block is wrapped in
``<!-- spreadsheet-import:proposal:KEY -->`` markers; the next run
strips its own prior blocks before evaluating, so unchanged data
produces no churn and updated spreadsheet rows refresh the block in
place.

**Renamed projects** are handled transparently: GitLab follows the
rename on GET, and the script uses the issue's numeric ``project_id``
for the subsequent PUT so write requests aren't affected by the
redirect-method downgrade GitLab applies on 301/302.

**Moved issues** (an issue physically moved to a different project,
which leaves a closed placeholder at the original iid with
``moved_to_id`` pointing to the new issue's numeric id) are
automatically followed via ``GET /api/v4/issues/<global-id>``. A
stderr line records the source → destination chain for each follow,
and the PUT lands on the live destination using its current
``project_id``. A chain of up to 5 hops is supported (in case an
issue was moved multiple times); deeper chains abort with an error.

Quick start::

    pip install -r requirements.txt openpyxl
    export GITLAB_TOKEN=<token-with-api-scope>

    # Eyeball one issue first
    python scripts/import_smartsheet.py \\
        --xlsx ESC_Risk_Register.xlsx --issue 555 --dry-run

    # Eyeball the first three matched issues
    python scripts/import_smartsheet.py \\
        --xlsx ESC_Risk_Register.xlsx --limit 3 --dry-run

    # Full run for real
    python scripts/import_smartsheet.py --xlsx ESC_Risk_Register.xlsx

SPDX-License-Identifier: GPL-3.0-or-later
Copyright (C) 2026 Ewan Douglas and contributors
"""

from __future__ import annotations

import argparse
import difflib
import hashlib
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

import requests

# openpyxl is imported lazily inside read_xlsx() so the rest of this
# module (merge_sections, parse_issue_url, the heading-normalisation
# helpers, etc.) can be imported and unit-tested without it.


# Mirrors build.py CANONICAL_SECTIONS, but with a *broader* synonym list
# so the importer also recognises legacy heading text.
# Note: legacy headings are *preserved* in-place; the canonical form is
# only used for proposal block labels and for matching/comparison.
# Order matters: missing sections are appended in this order.
SECTIONS: list[tuple[str, str, list[str]]] = [
    ("risk_description", "Risk Description",
     ["risk description", "description", "summary"]),
    ("notes", "Notes",
     # Old shapes that should migrate to "Notes":
     ["notes", "action plan / notes", "action plan/ notes",
      "action plan/notes", "action plan", "action"]),
    ("mitigation_plan", "Mitigation Plan",
     ["mitigation plan", "risk mitigation planning", "risk mitigation",
      "mitigation", "plan", "planning"]),
]

# Spreadsheet column header (normalised lowercase) → canonical section key.
COL_HEADERS: dict[str, str] = {
    "risk description": "risk_description",
    "action plan / notes": "notes",
    "action plan/ notes": "notes",
    "action plan/notes": "notes",
    "notes": "notes",
    "risk mitigation planning": "mitigation_plan",
    "mitigation plan": "mitigation_plan",
}

HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*$", re.MULTILINE)
ISSUE_URL_RE = re.compile(
    r"(?P<server>https?://[^/\s]+)"
    r"/(?P<path>.+?)"
    r"/-/(?:issues|work_items)/(?P<iid>\d+)"
)


# Proposal-block markers come in two forms:
#   - new format with a per-row source_id:
#       <!-- spreadsheet-import:proposal:KEY:SID -->
#   - legacy format without sid (written by earlier versions of this
#     script). Stripped on every run so users see a clean upgrade path.
NEW_PROPOSAL_RE = re.compile(
    r"<!-- spreadsheet-import:proposal:(?P<key>[\w-]+):[\w-]+ -->"
    r".*?"
    r"<!-- /spreadsheet-import:proposal:(?P=key):[\w-]+ -->\n?",
    re.DOTALL,
)
LEGACY_PROPOSAL_RE = re.compile(
    r"<!-- spreadsheet-import:proposal:(?P<key>[\w-]+) -->"
    r".*?"
    r"<!-- /spreadsheet-import:proposal:(?P=key) -->\n?",
    re.DOTALL,
)
# Back-compat alias for code/tests that imported PROPOSAL_RE.
PROPOSAL_RE = LEGACY_PROPOSAL_RE

_SAFE_SID_RE = re.compile(r"[^a-zA-Z0-9_-]+")


def _slugify_source_id(s: str | None) -> str:
    """Make a source identifier safe to embed in an HTML comment marker.

    Letters / digits / dash / underscore are preserved; anything else
    collapses to a dash. Empty input falls back to ``default`` so the
    marker shape stays stable.
    """
    if not s:
        return "default"
    slug = _SAFE_SID_RE.sub("-", str(s)).strip("-").lower()
    return slug or "default"


def _hash_sections(sections: dict[str, str]) -> str:
    """Stable 8-char identifier derived from a row's section content.

    Used as the source_id when the spreadsheet has no Unique Risk ID
    column. Same content → same hash → idempotent re-runs even without
    an explicit ID. Different content gets a different hash, so rows
    with distinct text coexist as separate proposal blocks rather than
    overwriting each other.
    """
    h = hashlib.sha256()
    for key in sorted(sections.keys()):
        h.update(key.encode("utf-8"))
        h.update(b"\0")
        h.update(sections[key].encode("utf-8"))
        h.update(b"\0")
    return h.hexdigest()[:8]


def _normalise_heading(h: str) -> str:
    s = h.strip().rstrip(":").lower()
    s = re.sub(r"\s*/\s*", " / ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _normalise_body(body: str) -> str:
    """Trim trailing whitespace per line and collapse blank-line runs so
    whitespace-only differences are treated as a match."""
    if not body:
        return ""
    out: list[str] = []
    prev_blank = False
    for line in body.splitlines():
        line = line.rstrip()
        if not line:
            if not prev_blank:
                out.append("")
            prev_blank = True
        else:
            out.append(line)
            prev_blank = False
    return "\n".join(out).strip()


_BLANK_RUN_RE = re.compile(r"\n{3,}")


def _normalise_blank_runs(text: str) -> str:
    """Collapse runs of 3+ newlines (≥2 blank lines) to a single blank
    line. Used after stripping a proposal block to clean up the empty
    space that would otherwise be left where the block used to sit."""
    return _BLANK_RUN_RE.sub("\n\n", text)


def _strip_proposals(text: str) -> str:
    """Remove ANY proposal block (any source_id, plus legacy no-sid).

    Used by tests and any caller that just wants a clean view of an
    issue's description. merge_sections uses a tighter strip that only
    removes blocks for the current row's source_id.
    """
    text = NEW_PROPOSAL_RE.sub("", text or "")
    text = LEGACY_PROPOSAL_RE.sub("", text)
    return _normalise_blank_runs(text)


def _strip_proposals_for_sid(text: str, sid: str) -> str:
    """Remove only the proposal blocks tagged with `sid`, plus any
    legacy (no-sid) blocks. Other rows' blocks survive untouched —
    that's what lets multiple spreadsheet rows targeting the same
    issue produce multiple coexisting proposal blocks.
    """
    sid_esc = re.escape(sid)
    pattern = re.compile(
        r"<!-- spreadsheet-import:proposal:(?P<key>[\w-]+):" + sid_esc + r" -->"
        r".*?"
        r"<!-- /spreadsheet-import:proposal:(?P=key):" + sid_esc + r" -->\n?",
        re.DOTALL,
    )
    text = pattern.sub("", text or "")
    text = LEGACY_PROPOSAL_RE.sub("", text)
    return _normalise_blank_runs(text)


def canonical_key(heading: str) -> str | None:
    norm = _normalise_heading(heading)
    for key, _, syns in SECTIONS:
        if norm in syns:
            return key
    return None


def _first_canonical_bodies(text: str) -> dict[str, str]:
    """Return {canonical_key: body_text} from the first canonical heading
    for each key encountered in `text`.

    Fallback for ``risk_description``: if no explicit canonical heading
    matches, the description's leading prose (text before the first
    markdown heading, or the whole body if it has no headings) is used.
    Mirrors :func:`build.parse_sections` so an issue whose body is just
    a bare risk statement — no ``## Risk Description`` header — is
    treated as already having that content, suppressing duplicate
    proposal blocks on re-import.
    """
    bodies: dict[str, str] = {}
    matches = list(HEADING_RE.finditer(text))
    for i, m in enumerate(matches):
        key = canonical_key(m.group(2))
        if not key or key in bodies:
            continue
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        bodies[key] = text[start:end].strip()
    if "risk_description" not in bodies:
        leading_end = matches[0].start() if matches else len(text)
        leading = text[:leading_end].strip()
        if leading:
            bodies["risk_description"] = leading
    return bodies


def _build_proposal_block(key: str, canonical_h: str, new_body: str,
                          existing_body: str, today: str,
                          source_label: str, source_id: str,
                          unique_id: str | None = None,
                          modification_date: str | None = None) -> str:
    """Render a markdown block proposing a new section value.

    Layout:
      * Heading for the proposed section. When the issue ALREADY has a
        canonical section with this name, the heading is preceded by a
        horizontal rule and tagged ``(imported from spreadsheet)`` so a
        reader can immediately tell the imported copy apart from the
        issue's own section above — e.g.::

            ---

            ### Mitigation Plan _(imported from spreadsheet)_

        When no such section exists yet, a plain ``### <name>`` heading is
        used (there's nothing to disambiguate from). Either way the
        heading is H3 so the issue's own H1/H2 section, if present, stays
        the first match for the dashboard's section parser; the decorated
        "imported" heading additionally no longer matches a canonical key
        at all, so it can never be mistaken for the authoritative section.
      * The new content.
      * If the issue had a prior version of this section, a unified diff
        in a ```diff code block (always visible — never inside <details>).
        Sections with no prior content omit the diff entirely.
      * Italic attribution footer: ``*(imported from <source>, on <date>)*``
        — when a Unique Risk ID is available it's woven in:
        ``*(imported from <source>, Unique Risk ID: <id>, on <date>)*``.

    The proposal's HTML-comment markers include the source_id so multiple
    spreadsheet rows targeting the same issue each get their own block.
    """
    sid = _slugify_source_id(source_id)
    parts: list[str] = [f"<!-- spreadsheet-import:proposal:{key}:{sid} -->"]
    if existing_body:
        # An existing section with this name sits above; set the imported
        # copy apart with a rule + an explicit "imported" tag. The blank
        # line before ``---`` keeps it a thematic break (an <hr>), not a
        # setext underline for the HTML-comment marker.
        parts.extend([
            "",
            "---",
            "",
            f"### {canonical_h} _(imported from spreadsheet)_",
        ])
    else:
        parts.append(f"### {canonical_h}")
    parts.extend([
        "",
        new_body.rstrip(),
    ])
    if existing_body:
        diff_lines = list(difflib.unified_diff(
            existing_body.splitlines(),
            new_body.splitlines(),
            fromfile="current",
            tofile="spreadsheet",
            lineterm="",
        ))
        diff_text = "\n".join(diff_lines) if diff_lines else (
            "(no textual diff — whitespace or formatting only)"
        )
        parts.extend([
            "",
            "```diff",
            diff_text,
            "```",
        ])
    attribution_bits = [f"imported from {source_label}"]
    if unique_id:
        attribution_bits.append(f"Unique Risk ID: {unique_id}")
    if modification_date:
        attribution_bits.append(f"Modification Date: {modification_date}")
    attribution_bits.append(f"on {today}")
    attribution = f"*({', '.join(attribution_bits)})*"
    parts.extend(["", attribution])
    parts.append(f"<!-- /spreadsheet-import:proposal:{key}:{sid} -->")
    return "\n".join(parts)


def merge_sections(existing: str | None, new_sections: dict[str, str],
                   today: str | None = None,
                   source_label: str = "spreadsheet",
                   source_id: str | None = None,
                   unique_id: str | None = None,
                   modification_date: str | None = None) -> str:
    """Insert a 'proposed update' block for each canonical section whose
    body in ``existing`` differs from the value in ``new_sections``.

    Never modifies existing content: the canonical sections (and any other
    headings or prose) in ``existing`` are preserved verbatim. New
    information from the spreadsheet is added as an **addendum directly
    after the matching existing section** (not dumped at the bottom),
    wrapped in HTML markers so re-runs can detect and refresh prior
    proposal blocks rather than duplicate them.

    Behavior per canonical section:

    - If the existing canonical section body matches the spreadsheet
      value (modulo trailing whitespace and blank-line runs), nothing
      is added for that section.
    - If they differ, a proposal block with the new content and a unified
      diff is inserted immediately after that section's existing body
      (before the next heading) so the addendum reads in context.
    - If no matching canonical heading exists in ``existing`` (a brand-new
      section, or the bare-prose risk-description fallback), the proposal
      block is appended at the end in canonical order.
    - Sections not present in ``new_sections`` are ignored.

    ``source_id`` tags each proposal block's HTML markers so multiple
    spreadsheet rows targeting the same issue produce coexisting blocks
    instead of overwriting each other. On re-run, only blocks tagged
    with the same source_id are stripped before re-emitting — other
    rows' blocks survive untouched. ``source_id`` defaults to a stable
    hash of ``new_sections`` when not provided, so multiple rows with
    distinct content still coexist even without a Unique Risk ID.

    ``unique_id``, when present, is rendered inline in the attribution
    line as ``Unique Risk ID: <id>`` for human readability.
    """
    if today is None:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if source_id is None:
        source_id = _hash_sections(new_sections)
    sid = _slugify_source_id(source_id)

    # Two views of the existing description:
    #   - `text`: the version we'll write back to GitLab. Strips this
    #     row's prior proposal blocks (so re-runs refresh in place) plus
    #     any legacy no-sid blocks (so an older format upgrades cleanly).
    #     Blocks from OTHER rows targeting the same issue stay put.
    #   - `clean`: every proposal block of any sid removed. Used only to
    #     identify the issue's canonical section bodies for the diff —
    #     we never want to diff against another row's H3 proposal.
    text = _strip_proposals_for_sid(existing or "", sid).rstrip()
    clean = _strip_proposals(existing or "")
    real_bodies = _first_canonical_bodies(clean)

    proposals_by_key: dict[str, str] = {}
    for key, canonical_h, _ in SECTIONS:
        if key not in new_sections:
            continue
        new_body = new_sections[key].strip()
        existing_body = real_bodies.get(key, "")
        if _normalise_body(existing_body) == _normalise_body(new_body):
            continue
        proposals_by_key[key] = _build_proposal_block(
            key, canonical_h, new_body, existing_body, today,
            source_label, source_id, unique_id, modification_date,
        )

    if not proposals_by_key:
        return (text + "\n") if text else ""
    merged = _insert_or_append_proposals(text, proposals_by_key)
    merged = _normalise_blank_runs(merged).rstrip()
    return merged + "\n"


def _proposal_spans(text: str) -> list[tuple[int, int]]:
    """Character spans of every proposal block (new-format and legacy) in
    `text`. Used to tell a *real* canonical heading apart from the
    ``### Risk Description`` heading that lives inside another row's
    proposal block, so insertion anchors on real sections only."""
    spans: list[tuple[int, int]] = []
    for rx in (NEW_PROPOSAL_RE, LEGACY_PROPOSAL_RE):
        for m in rx.finditer(text):
            spans.append((m.start(), m.end()))
    return spans


def _insert_or_append_proposals(text: str,
                                proposals_by_key: dict[str, str]) -> str:
    """Place each proposal block right after the *existing* canonical
    section it refers to, rather than dumping them all at the end.

    For a section that already has a real heading in ``text`` (e.g.
    ``## Notes``), the proposal is inserted just before the next real
    canonical/other heading — i.e. immediately after that section's body,
    as an addendum. Sections that don't yet exist in ``text`` (and the
    bare-prose fallback) are appended at the end in canonical order.

    "Real" headings exclude any heading that lives inside another row's
    proposal block, so addenda group after the section they belong to and
    re-runs stay stable.
    """
    if not text:
        blocks = [proposals_by_key[k] for k, _, _ in SECTIONS
                  if k in proposals_by_key]
        return "\n\n".join(blocks)

    spans = _proposal_spans(text)

    def in_proposal(pos: int) -> bool:
        return any(s <= pos < e for s, e in spans)

    real_headings = [m for m in HEADING_RE.finditer(text)
                     if not in_proposal(m.start())]
    # section_end_for_key: where the first real occurrence of each canonical
    # section's content stops = start of the *next real heading*, or EOF.
    # Using the next real heading (not the next proposal block) means a
    # re-run appends this row's refreshed block after any sibling addenda
    # already sitting under the section, keeping order stable.
    section_end_for_key: dict[str, int] = {}
    for idx, m in enumerate(real_headings):
        k = canonical_key(m.group(2))
        if k and k not in section_end_for_key:
            section_end_for_key[k] = (
                real_headings[idx + 1].start()
                if idx + 1 < len(real_headings) else len(text)
            )

    inserts: list[tuple[int, str]] = []
    appended: list[str] = []
    for key, _, _ in SECTIONS:
        if key not in proposals_by_key:
            continue
        block = proposals_by_key[key]
        if key in section_end_for_key:
            inserts.append((section_end_for_key[key], block))
        else:
            appended.append(block)

    # Apply insertions highest-offset-first so lower offsets stay valid.
    result = text
    for offset, block in sorted(inserts, key=lambda t: t[0], reverse=True):
        before = result[:offset].rstrip()
        after = result[offset:]
        if after.strip():
            result = f"{before}\n\n{block}\n\n{after.lstrip(chr(10))}"
        else:
            result = f"{before}\n\n{block}"

    if appended:
        result = result.rstrip() + "\n\n" + "\n\n".join(appended)
    return result


def parse_issue_url(url: str) -> dict | None:
    m = ISSUE_URL_RE.search(url.strip())
    if not m:
        return None
    return {
        "server": m.group("server"),
        "project_path": m.group("path"),
        "iid": int(m.group("iid")),
    }


def _project_path_from_web_url(web_url: str | None) -> str | None:
    """Extract the canonical project path from an issue's web_url. Returns
    the path GitLab itself uses (after any rename/move) so a PUT lands
    on the right project without depending on a redirect."""
    if not web_url:
        return None
    parsed = parse_issue_url(web_url)
    return parsed["project_path"] if parsed else None


def read_xlsx(path: Path, sheet: str | None) -> list[dict]:
    import openpyxl  # lazy: only needed for the main CLI flow
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    iter_rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(iter_rows)
    except StopIteration:
        return []
    headers = [str(c or "").strip() for c in header_row]
    rows: list[dict] = []
    for raw in iter_rows:
        if not any(raw):
            continue
        rows.append({h: v for h, v in zip(headers, raw) if h})
    return rows


def extract_sections(row: dict) -> dict[str, str]:
    out: dict[str, str] = {}
    for header, value in row.items():
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        key = COL_HEADERS.get(header.lower().strip())
        if key and key not in out:  # first column wins if duplicates
            out[key] = text
    return out


def find_link(row: dict) -> str | None:
    for k, v in row.items():
        if not v:
            continue
        kl = k.lower()
        if "gitlab" in kl and "link" in kl:
            return str(v)
    return None


_UNIQUE_ID_HEADERS = {"unique risk id", "risk id", "risk_id", "risk #"}
_MOD_DATE_HEADERS = {
    "modification date", "modified date", "last modified",
    "modified", "updated", "date modified",
}


def _find_row_value(row: dict, accepted: set[str]) -> str | None:
    """Return the first non-empty value whose normalised header is in
    `accepted`. Date-typed Excel cells become ``YYYY-MM-DD`` strings."""
    for header, value in row.items():
        if value is None:
            continue
        if header is None:
            continue
        key = " ".join(str(header).strip().lower().split())
        if key in accepted:
            if hasattr(value, "strftime"):
                # openpyxl returns datetime objects for date-typed cells.
                return value.strftime("%Y-%m-%d")
            text = str(value).strip()
            if text:
                return text
    return None


def find_unique_id(row: dict) -> str | None:
    """Look for a 'Unique Risk ID' (or close synonym) column in a row."""
    return _find_row_value(row, _UNIQUE_ID_HEADERS)


def find_modification_date(row: dict) -> str | None:
    """Look for a Modification Date column in a row. Returns ``YYYY-MM-DD``
    for datetime-typed cells, otherwise the raw stripped string."""
    return _find_row_value(row, _MOD_DATE_HEADERS)


def get_issue(session: requests.Session, server: str, project: str, iid: int) -> dict:
    url = f"{server}/api/v4/projects/{quote(project, safe='')}/issues/{iid}"
    r = session.get(url, timeout=30)
    r.raise_for_status()
    return r.json()


def get_issue_by_id(session: requests.Session, server: str, issue_id: int) -> dict:
    """Fetch a single issue by its global numeric id (not the per-project iid).

    Used to follow ``moved_to_id`` to the destination issue without
    needing to know the destination project's path or iid up front.
    """
    url = f"{server}/api/v4/issues/{issue_id}"
    r = session.get(url, timeout=30)
    r.raise_for_status()
    return r.json()


def put_description(session: requests.Session, server: str,
                    project: str | int, iid: int, description: str) -> dict:
    """PUT a new description for an issue.

    `project` may be either the URL-encoded full path
    (``group/subgroup/project``) or the numeric project ID. Numeric IDs
    are recommended after a GET so we sidestep rename redirects — GitLab
    follows 301/302 on GET but not on PUT, which produces a 405.
    """
    project_segment = str(project) if isinstance(project, int) else quote(project, safe="")
    url = f"{server}/api/v4/projects/{project_segment}/issues/{iid}"
    r = session.put(url, json={"description": description}, timeout=30)
    r.raise_for_status()
    return r.json()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--xlsx", required=True, type=Path,
                    help="Smartsheets .xlsx export")
    ap.add_argument("--token",
                    help="GitLab access token with api scope "
                         "(falls back to GITLAB_TOKEN env var)")
    ap.add_argument("--token-env", default="GITLAB_TOKEN",
                    help="Env var name holding the token (default: GITLAB_TOKEN)")
    ap.add_argument("--server",
                    help="Override GitLab base URL (default: parse from each row's URL)")
    ap.add_argument("--sheet",
                    help="Worksheet name (default: active sheet)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print a unified diff per changed issue; do not write")
    ap.add_argument("--print-markdown", action="store_true",
                    help="Print the full proposed issue description as clean "
                         "markdown (pasteable into GitLab's preview) instead "
                         "of a unified diff. Implies --dry-run (never writes). "
                         "Pair with --issue to isolate one issue.")
    ap.add_argument("--limit", type=int, default=0,
                    help="Process at most N matched rows (0 = no limit)")
    ap.add_argument("--issue", type=int,
                    help="Process only the issue with this iid (debugging)")
    ap.add_argument("--backup", type=Path,
                    default=Path(f"backup-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.jsonl"),
                    help="Append original issue descriptions to this file before writing")
    args = ap.parse_args()
    # --print-markdown is a read-only preview mode: force dry-run so no
    # write path can ever fire, regardless of how the output branch below
    # is structured.
    if args.print_markdown:
        args.dry_run = True

    token = args.token or os.environ.get(args.token_env)
    if not token:
        print(f"error: set --token or ${args.token_env}", file=sys.stderr)
        return 2

    rows = read_xlsx(args.xlsx, args.sheet)
    print(f"Loaded {len(rows)} rows from {args.xlsx}"
          + (f" (sheet: {args.sheet})" if args.sheet else ""))

    session = requests.Session()
    session.headers["Authorization"] = f"Bearer {token}"

    stats = dict(total=0, no_link=0, no_sections=0, no_change=0,
                 followed_move=0, would_update=0, updated=0, errors=0)
    # Per-row failure log keyed by Unique Risk ID (or fallback row label).
    # Each entry is (row_tag, reason). Printed at the end so users can see
    # at a glance which spreadsheet rows didn't make it to GitLab and why,
    # without having to scrub the per-row error stream.
    failures: list[tuple[str, str]] = []

    processed = 0
    with args.backup.open("a") as backup_f:
        for row_idx, row in enumerate(rows, start=2):
            # row_idx starts at 2 to match the xlsx row number the user sees
            # in their spreadsheet (row 1 is the header).
            stats["total"] += 1
            unique_id = find_unique_id(row)
            row_tag = unique_id or f"row {row_idx}"
            link = find_link(row)
            if not link:
                print(f"  - {row_tag}: no GitLab Link column populated",
                      file=sys.stderr)
                failures.append((row_tag, "no GitLab Link in spreadsheet"))
                stats["no_link"] += 1
                continue
            parsed = parse_issue_url(str(link))
            if not parsed:
                print(f"  ! {row_tag}: unparseable GitLab Link: {link!r}",
                      file=sys.stderr)
                failures.append((row_tag, f"unparseable GitLab Link: {link!r}"))
                stats["no_link"] += 1
                continue
            if args.issue and parsed["iid"] != args.issue:
                continue
            new_sections = extract_sections(row)
            if not new_sections:
                print(f"  - {row_tag}: no recognized section columns "
                      f"(Risk Description / Notes / Mitigation)",
                      file=sys.stderr)
                failures.append((row_tag, "no recognized section columns"))
                stats["no_sections"] += 1
                continue

            server = args.server or parsed["server"]
            project = parsed["project_path"]
            iid = parsed["iid"]
            try:
                issue = get_issue(session, server, project, iid)
            except requests.HTTPError as e:
                print(f"  ! {row_tag} → GET {project}#{iid}: {e}",
                      file=sys.stderr)
                failures.append(
                    (row_tag, f"GET {project}#{iid} failed: {e}")
                )
                stats["errors"] += 1
                continue

            # If the issue was moved (the spreadsheet URL points at a
            # closed placeholder), follow the chain to the live destination
            # so the write lands on the right issue. Renames are handled
            # transparently elsewhere via numeric project_id on the PUT —
            # this branch only fires for actual project-to-project moves.
            move_chain: list[str] = []
            move_followed = False
            move_failure_reason: str | None = None
            while issue.get("moved_to_id"):
                if len(move_chain) >= 5:
                    print(
                        f"  ! {row_tag} → {project}#{iid}: move chain too deep "
                        f"(>{len(move_chain)} hops). Aborting follow.",
                        file=sys.stderr,
                    )
                    move_failure_reason = (
                        f"move chain from {move_chain[0]} too deep "
                        f"(>{len(move_chain)} hops)"
                    )
                    issue = None
                    break
                move_chain.append(f"{project}#{iid}")
                target_id = issue["moved_to_id"]
                try:
                    issue = get_issue_by_id(session, server, target_id)
                except requests.HTTPError as e:
                    print(
                        f"  ! {row_tag} → following move from "
                        f"{move_chain[0]} (moved_to_id={target_id}): {e}\n"
                        f"    Hint: the token needs read access to the "
                        f"destination project, or update the spreadsheet's "
                        f"GitLab Link to the destination URL directly.",
                        file=sys.stderr,
                    )
                    move_failure_reason = (
                        f"move from {move_chain[0]} → "
                        f"moved_to_id={target_id}: {e}"
                    )
                    issue = None
                    break
                iid = issue.get("iid", iid)
                project = (
                    _project_path_from_web_url(issue.get("web_url"))
                    or project
                )
                move_followed = True
            if issue is None:
                failures.append((row_tag, move_failure_reason or "move-follow failed"))
                stats["errors"] += 1
                continue
            if move_followed:
                print(
                    f"  > {row_tag}: followed move chain: "
                    f"{' → '.join(move_chain + [f'{project}#{iid}'])}",
                    file=sys.stderr,
                )
                stats["followed_move"] += 1

            current = issue.get("description") or ""
            backup_f.write(json.dumps({
                "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "server": server, "project": project, "iid": iid,
                "description": current,
            }) + "\n")

            modification_date = find_modification_date(row)
            new_desc = merge_sections(
                current, new_sections,
                source_label=args.xlsx.name,
                source_id=unique_id,  # falls back to row-hash when None
                unique_id=unique_id,
                modification_date=modification_date,
            )
            if new_desc == current:
                stats["no_change"] += 1
                continue

            # Prefer the canonical project path or numeric id from the GET
            # response — the spreadsheet URL may point at an old path that
            # GitLab now 301-redirects, and that redirect doesn't preserve
            # the method on PUT (results in 405 Method Not Allowed).
            project_for_put: str | int = (
                issue.get("project_id")
                or _project_path_from_web_url(issue.get("web_url"))
                or project
            )

            display_project = (
                _project_path_from_web_url(issue.get("web_url")) or project
            )
            print(f"\n--- {row_tag} → {display_project}#{iid} ---  "
                  f"({issue.get('web_url') or link})")
            if args.print_markdown:
                # Emit the full proposed description verbatim so it can be
                # pasted into the issue and previewed. The delimiter is an
                # HTML comment, which is invisible in GitLab's rendered
                # markdown — so even copying the whole block (separator
                # included) previews cleanly. Pair with --issue to get a
                # single block on stdout.
                sys.stdout.write(
                    f"<!-- ===== {row_tag} → {display_project}#{iid} "
                    f"(paste below into the issue, then Preview) ===== -->\n"
                )
                sys.stdout.write(
                    new_desc if new_desc.endswith("\n") else new_desc + "\n"
                )
                sys.stdout.write("\n")
                stats["would_update"] += 1
            elif args.dry_run:
                diff = "".join(difflib.unified_diff(
                    current.splitlines(keepends=True),
                    new_desc.splitlines(keepends=True),
                    fromfile="current", tofile="new", lineterm="",
                ))
                sys.stdout.write(diff or "(rewritten with no textual diff)\n")
                stats["would_update"] += 1
            else:
                try:
                    put_description(session, server, project_for_put,
                                    iid, new_desc)
                    print("  ✓ updated")
                    stats["updated"] += 1
                except requests.HTTPError as e:
                    print(f"  ! {row_tag} → PUT {display_project}#{iid}: {e}",
                          file=sys.stderr)
                    failures.append(
                        (row_tag, f"PUT {display_project}#{iid} failed: {e}")
                    )
                    stats["errors"] += 1

            processed += 1
            if args.limit and processed >= args.limit:
                break

    print("\n=== Summary ===")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    if failures:
        print("\n=== Rows that did not update ===")
        # Sorted by row tag for predictable scanning; the user's spreadsheet
        # is usually sorted by Unique Risk ID, so this matches their mental
        # model better than insertion order.
        for tag, reason in sorted(failures, key=lambda t: t[0]):
            print(f"  - {tag}: {reason}")
    print(f"\nBackup of original descriptions: {args.backup}")
    return 0 if stats["errors"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
